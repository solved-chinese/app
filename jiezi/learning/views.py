from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from learning.models import Character, Radical
from accounts.models import UserCharacter
from django.contrib.auth.decorators import user_passes_test
import openpyxl, random, math, datetime

MIN_LEARN_REVIEW_INTERVAL=30

def index(request):
    return render(request, 'index.html');

def about_us(request):
    return  render(request, 'about_us.html');

@login_required()
def start_learning (request, minutes_to_learn):
    request.session.cycle_key()
    request.session['end_time']=timezone.now()+datetime.timedelta(minutes=minutes_to_learn)
    request.session['minutes_to_learn']=minutes_to_learn
    request.session['current_stage']=0
    request.session['uc_pk']= request.user.user_characters.first().pk
    return redirect(f'/learning/status{request.session.session_key}')

#1 learn 2 pron 3 meaning 4 tpron 5 tmeaning 6 relearn 0 decideL/R
@login_required()
def user_status (request, session_key):
    if request.session.session_key != session_key:
        return redirect('/404')
    if request.POST.get('choice') is not None:
        return review_interface(request)
    if timezone.now()>request.session['end_time']:
        return render(request, 'simple_response.html', {'content':f'Congratulations! You have finished studying for {request.session["minutes_to_learn"]} minutes. Take a break :).'})
    try:
        current_stage=request.session['current_stage']=transition_stage(request)
    except Exception as e:
        return render(request, 'simple_response.html', {'content':str(e)+'<br>You have finished everything! Please come back later to review or add some more to your learning stack!'})
    uc=UserCharacter.objects.get(pk=request.session['uc_pk'])
    if current_stage == 1 or current_stage == 6:
        return view_character(request, uc.character.pk, is_view=False)
    elif current_stage == 2 or current_stage==4:
        list = []
        for user_character in request.user.user_characters.all():
            list.append(user_character.character.pinyin)
        if len(list) < 4:
            list.extend(['hǎn', 'dā', 'zhé'])
        return review_interface(request, list, uc.character.pinyin,
                                f'"<span style="color:  #B62C25">{uc.character.chinese}</span>" is pronounced ___:')
    elif current_stage==3 or current_stage==5:
        list = []
        for user_character in request.user.user_characters.all():
            list.append(user_character.character.definition_1)
        if len(list) < 4:
            list.extend(['powerful','meaningless','interesting'])
        return review_interface(request, list, uc.character.definition_1,
                                f'"<span style="color:  #B62C25">{uc.character.chinese}</span>" means ___:')

def transition_stage(request):
    current_stage = request.session['current_stage']
    # print('stage was'+str(current_stage)+'correct was'+str(request.session['correct']))
    if current_stage==1 or current_stage==2 or current_stage==4:
        current_stage+=1
    elif current_stage==5:
        uc=UserCharacter.objects.get(pk=request.session['uc_pk'])
        if request.session['correct']==True:
            current_stage=0
            uc.update(1)
        else:
            current_stage=6
            uc.update(0)
    elif current_stage==6:
        current_stage=0
    elif current_stage==3:
        current_stage=0
        UserCharacter.objects.get(pk=request.session['uc_pk']).update(-1)
    if current_stage==0:
        RV = is_learn(request.user)
        if RV[0]:
            current_stage = 1
            request.session['uc_pk']=RV[1]
        else:
            current_stage=4
            request.session['uc_pk']=RV[1]
    if current_stage == 4:
        request.session['correct'] = True
    print('stage now' + str(current_stage))
    return current_stage

def is_learn(user):
    to_learn = user.user_characters.filter(times_learned=0).first()
    to_review = user.user_characters.filter(time_last_learned__lt=timezone.now() - datetime.timedelta(seconds=MIN_LEARN_REVIEW_INTERVAL),times_learned__gte=1).first()
    if to_review is None and to_learn is None:
        raise Exception('NOTHING_TO_LEARN_OR_REVIEW')
    if to_learn is None:
        if to_review.times_learned>2:
            raise Exception('MASTERED_EVERYTHING')
        else:
            return (False, to_review.pk)
    if to_review is None:
        return (True, to_learn.pk)
    review_prob = (5.2-math.log(2*to_review.interval))/7
    review_prob = min(0.7, max(0.01, review_prob))
    # the more familiar the user is with all characters, the more likely it is for him to learn a new one
    print("cha:"+str(to_review)+'prob:' + str(review_prob))
    if random.random() < review_prob:
        return (False, to_review.pk)
    return (True, to_learn.pk)


@login_required
def view_character(request, character_pk, is_view=True, pure=False):
    character = Character.objects.get(pk=character_pk)
    dict = {'character': character, 'radical_1': Radical.objects.get(pk=character.mnemonic_1)}
    try:
        radical_2 = Radical.objects.get(pk=character.mnemonic_2)
        dict['radical_2'] = radical_2
    except:
        dict['radical_2'] = None
    try:
        radical_3 = Radical.objects.get(pk=character.mnemonic_3)
        dict['radical_3'] = radical_3
    except:
        dict['radical_3'] = None
    dict['is_view']=is_view
    if pure==True:
        return render(request, 'learning/learning_character_pure.html', dict)
    return render(request, 'learning/learning_character.html', dict)

@login_required
def review_interface(request, list=[], ans='', question=''):
    if request.POST.get('choice') is not None:
        uc=UserCharacter.objects.get(pk=request.session['uc_pk'])
        current_stage=request.session['current_stage']
        choice = int(request.POST.get('choice'))
        correct = request.session['ans_index']
        if choice == correct:
            incorrect = -1
        else:
            incorrect = choice
            if current_stage>3:
                print('set correct to false')
                request.session['correct']=False

        choices = [request.session['choice1'], request.session['choice2'],request.session['choice3'],request.session['choice4']]
        return render(request, 'learning/review_interface.html',
                      {'choices': choices, 'question': request.session['question'], 'correct': correct, 'incorrect': incorrect})
    else:
        choices = random.sample(list, 4)
        ans_index = -1
        for i in range(0, 4):
            if choices[i] == ans:
                ans_index = i
        if ans_index == -1:
            ans_index = random.randrange(4)
            choices[ans_index] = ans
        try:
            request.user.quiz.delete()
        except:
            pass
        request.session['choice1']=choices[0]
        request.session['choice2'] = choices[1]
        request.session['choice3'] = choices[2]
        request.session['choice4'] = choices[3]
        request.session['question'] = question
        request.session['ans_index'] = ans_index
        return render(request, 'learning/review_interface.html',
                      {'choices': choices, 'question': question, 'correct': -1, 'incorrect': -1})


@user_passes_test(lambda u: u.is_staff)
def load_radical(request):
    if "GET" == request.method:
        return render(request, 'learning/load_excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Radicals"]
        list = []
        n = worksheet.max_row
        m = worksheet.max_column
        for i in range(2, n + 1):
            pk = worksheet.cell(i, 1).value
            if pk is None:
                continue
            pk = int(pk)
            chinese = worksheet.cell(i, 2).value
            pinyin = worksheet.cell(i, 3).value
            definition = worksheet.cell(i, 4).value
            mnemonic_explanation = worksheet.cell(i, 5).value
            is_phonetic = True if str(worksheet.cell(i, 6).value) == '1' else False
            is_semantic = True if str(worksheet.cell(i, 7).value) == '1' else False
            message = ''
            try:
                if Radical.objects.filter(pk=pk).exists():
                    message += 'update R%04d ' % pk
                    Radical.objects.filter(pk=pk).update(chinese=chinese, pinyin=pinyin, definition=definition,
                                                         mnemonic_explanation=mnemonic_explanation,
                                                         is_phonetic=is_phonetic, is_semantic=is_semantic)
                else:
                    message += 'create R%04d ' % pk
                    Radical.objects.create(jiezi_id=pk, chinese=chinese, pinyin=pinyin, definition=definition,
                                           mnemonic_explanation=mnemonic_explanation, is_phonetic=is_phonetic,
                                           is_semantic=is_semantic)
            except Exception as e:
                message += str(e)
            else:
                message += 'success'
            list.append(message)
        return render(request, 'learning/load_excel.html', {"list": list})


@user_passes_test(lambda u: u.is_staff)
def load_character(request):
    if "GET" == request.method:
        return render(request, 'learning/load_excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Characters"]
        list = []
        n = worksheet.max_row
        m = worksheet.max_column
        for i in range(2, n + 1):
            pk = worksheet.cell(i, 2).value
            if pk is None or pk=='':
                continue
            pk = int(pk)
            chinese = worksheet.cell(i, 3).value
            pinyin = worksheet.cell(i, 4).value
            definition_1 = worksheet.cell(i, 5).value
            definition_2 = worksheet.cell(i, 6).value
            explanation_2 = worksheet.cell(i, 7).value
            definition_3 = worksheet.cell(i, 8).value
            explanation_3 = worksheet.cell(i, 9).value
            r1 = worksheet.cell(i, 10).value
            r2 = worksheet.cell(i, 11).value
            if r2 == '': r2=None
            r3 = worksheet.cell(i, 12).value
            if r3 == '': r3=None
            mnemonic_explanation = worksheet.cell(i, 13).value
            example1_word = worksheet.cell(i, 14).value
            example1_pinyin = worksheet.cell(i, 15).value
            example1_character = worksheet.cell(i, 16).value
            example1_meaning = worksheet.cell(i, 17).value
            example2_word = worksheet.cell(i, 18).value
            example2_pinyin = worksheet.cell(i, 19).value
            example2_character = worksheet.cell(i, 20).value
            example2_meaning = worksheet.cell(i, 21).value
            message = ''
            try:
                if Character.objects.filter(pk=pk).exists():
                    message += 'update C%04d: ' % pk
                    Character.objects.filter(pk=pk).update(chinese=chinese, pinyin=pinyin, definition_1=definition_1,
                                                           definition_2=definition_2, explanation_2=explanation_2,
                                                           explanation_3=explanation_3, definition_3=definition_3,
                                                           mnemonic_1=r1, mnemonic_2=r2, mnemonic_3=r3,
                                                           mnemonic_explanation=mnemonic_explanation,
                                                           example_1_word=example1_word,
                                                           example_1_pinyin=example1_pinyin,
                                                           example_1_character=example1_character,
                                                           example_1_meaning=example1_meaning,
                                                           example_2_word=example2_word,
                                                           example_2_pinyin=example2_pinyin,
                                                           example_2_character=example2_character,
                                                           example_2_meaning=example2_meaning)
                else:
                    message += 'create C%04d: ' % pk
                    Character.objects.create(jiezi_id=pk, chinese=chinese, pinyin=pinyin, definition_1=definition_1,
                                             definition_2=definition_2, explanation_2=explanation_2,
                                             explanation_3=explanation_3, definition_3=definition_3, mnemonic_1=r1,
                                             mnemonic_2=r2, mnemonic_3=r3, mnemonic_explanation=mnemonic_explanation,
                                             example_1_word=example1_word, example_1_pinyin=example1_pinyin,
                                             example_1_character=example1_character, example_1_meaning=example1_meaning,
                                             example_2_word=example2_word, example_2_pinyin=example2_pinyin,
                                             example_2_character=example2_character, example_2_meaning=example2_meaning)
            except Exception as e:
                message += str(e)
            else:
                message += 'success'
            list.append(message)
        return render(request, 'learning/load_excel.html', {"list": list})

def test(request):
    return render(request,'accounts/manage.html')
